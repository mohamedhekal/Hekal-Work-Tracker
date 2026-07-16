"""Daily Excel backup/sync alongside JSON storage."""

from __future__ import annotations

import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

from core import (
    HourAdjustment,
    Session,
    Settings,
    calculate_day_summary,
    calculate_period_summary,
    period_bounds_for_date,
)
from storage import DATA_DIR, dump_state, load_payload, save_data

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover - runtime dependency check
    Workbook = None  # type: ignore[misc, assignment]
    load_workbook = None  # type: ignore[misc, assignment]
    Font = None  # type: ignore[misc, assignment]


EXCEL_FILE = DATA_DIR / "work_hours_backup.xlsx"

SHEET_SETTINGS = "الإعدادات"
SHEET_SESSIONS = "الجلسات"
SHEET_ADJUSTMENTS = "التعديلات"
SHEET_DAILY = "ملخص_يومي"
SHEET_MONTHLY = "ملخص_شهري"
SHEET_META = "معلومات"


class ExcelSyncError(Exception):
    pass


def _require_openpyxl() -> None:
    if Workbook is None or load_workbook is None:
        raise ExcelSyncError("مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl")


def should_sync_today(last_sync_date: str | None, today: date | None = None) -> bool:
    ref = today or date.today()
    return last_sync_date != ref.isoformat()


def export_to_excel(
    settings: Settings,
    sessions: list[Session],
    active_session_id: str | None,
    *,
    adjustments: list[HourAdjustment] | None = None,
    now: datetime | None = None,
    path: Path | None = None,
) -> Path:
    _require_openpyxl()
    assert Workbook is not None and Font is not None

    sync_time = now or datetime.now()
    adj_list = list(adjustments or ())
    target = path or EXCEL_FILE
    wb = Workbook()

    ws_settings = wb.active
    ws_settings.title = SHEET_SETTINGS
    ws_settings.append(["المفتاح", "القيمة"])
    for key, value in (
        ("month_start_day", settings.month_start_day),
        ("standard_daily_hours", settings.standard_daily_hours),
        ("standard_hourly_rate", settings.standard_hourly_rate),
        ("overtime_hourly_rate", settings.overtime_hourly_rate),
    ):
        ws_settings.append([key, value])

    ws_sessions = wb.create_sheet(SHEET_SESSIONS)
    ws_sessions.append(["id", "start", "end", "duration_hours", "status"])
    for session in sorted(sessions, key=lambda s: s.start):
        duration = session.duration_seconds(now=sync_time) / 3600.0
        status = "نشطة" if session.end is None else "منتهية"
        ws_sessions.append(
            [
                session.id,
                session.start.isoformat(timespec="seconds"),
                session.end.isoformat(timespec="seconds") if session.end else "",
                round(duration, 4),
                status,
            ]
        )

    ws_adj = wb.create_sheet(SHEET_ADJUSTMENTS)
    ws_adj.append(["id", "at", "hours_delta", "reason", "created_at", "type", "hour_type"])
    for adj in sorted(adj_list, key=lambda a: a.at):
        ws_adj.append(
            [
                adj.id,
                adj.at.isoformat(timespec="seconds"),
                round(adj.hours_delta, 4),
                adj.reason,
                adj.created_at.isoformat(timespec="seconds"),
                "إضافة" if adj.hours_delta > 0 else "خصم",
                adj.hour_type,
            ]
        )

    ws_daily = wb.create_sheet(SHEET_DAILY)
    ws_daily.append(
        [
            "day",
            "total_hours",
            "standard_hours",
            "overtime_hours",
            "earnings",
            "session_count",
            "adjustment_hours",
        ]
    )
    day_range = _data_day_range(sessions, adj_list, sync_time)
    if day_range is not None:
        first_day, last_day = day_range
        current = first_day
        while current <= last_day:
            summary = calculate_day_summary(
                current, sessions, settings, now=sync_time, adjustments=adj_list
            )
            if summary.total_hours > 0 or summary.adjustment_hours != 0 or summary.session_count > 0:
                ws_daily.append(
                    [
                        summary.day.isoformat(),
                        round(summary.total_hours, 4),
                        round(summary.standard_hours, 4),
                        round(summary.overtime_hours, 4),
                        round(summary.earnings, 2),
                        summary.session_count,
                        round(summary.adjustment_hours, 4),
                    ]
                )
            current += timedelta(days=1)

    ws_monthly = wb.create_sheet(SHEET_MONTHLY)
    ws_monthly.append(
        [
            "period_label",
            "start_date",
            "end_date",
            "total_hours",
            "standard_hours",
            "overtime_hours",
            "earnings",
        ]
    )
    seen_periods: set[str] = set()
    period_refs: list[date] = [s.start.date() for s in sessions]
    period_refs.extend(a.at.date() for a in adj_list)
    for ref_day in period_refs:
        period_start, _period_end, label = period_bounds_for_date(
            ref_day, settings.month_start_day
        )
        if label in seen_periods:
            continue
        seen_periods.add(label)
        period = calculate_period_summary(
            sessions, settings, period_start, now=sync_time, adjustments=adj_list
        )
        ws_monthly.append(
            [
                period.label,
                period.start.isoformat(),
                period.end.isoformat(),
                round(period.total_hours, 4),
                round(period.standard_hours, 4),
                round(period.overtime_hours, 4),
                round(period.earnings, 2),
            ]
        )

    ws_meta = wb.create_sheet(SHEET_META)
    ws_meta.append(["key", "value"])
    ws_meta.append(["last_sync_at", sync_time.isoformat(timespec="seconds")])
    ws_meta.append(["last_sync_date", sync_time.date().isoformat()])
    ws_meta.append(["excel_file", str(target)])
    ws_meta.append(["json_primary", "true"])
    ws_meta.append(["session_count", len(sessions)])
    ws_meta.append(["adjustment_count", len(adj_list)])
    ws_meta.append(["active_session_id", active_session_id or ""])

    for sheet in wb.worksheets:
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True) if sheet.max_column >= 2 else Font(bold=True)

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def _data_day_range(
    sessions: list[Session],
    adjustments: list[HourAdjustment],
    sync_time: datetime,
) -> tuple[date, date] | None:
    days: list[date] = []
    for session in sessions:
        days.append(session.start.date())
        days.append((session.end or sync_time).date())
    for adj in adjustments:
        days.append(adj.at.date())
    if not days:
        return None
    return min(days), max(days)


def import_from_excel(
    path: Path | None = None,
) -> tuple[Settings, list[Session], list[HourAdjustment], str | None]:
    _require_openpyxl()
    assert load_workbook is not None

    excel_path = path or EXCEL_FILE
    if not excel_path.exists():
        raise ExcelSyncError(f"ملف Excel غير موجود: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    if SHEET_SETTINGS not in wb.sheetnames:
        raise ExcelSyncError(f"ورقة '{SHEET_SETTINGS}' غير موجودة في ملف Excel")

    settings_map: dict[str, object] = {}
    for row in wb[SHEET_SETTINGS].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        settings_map[str(row[0])] = row[1]

    settings = Settings(
        month_start_day=int(settings_map.get("month_start_day", 17)),
        standard_daily_hours=float(settings_map.get("standard_daily_hours", 4)),
        standard_hourly_rate=float(settings_map.get("standard_hourly_rate", 10)),
        overtime_hourly_rate=float(settings_map.get("overtime_hourly_rate", 35)),
    )
    settings.validate()

    sessions: list[Session] = []
    active_session_id: str | None = None

    if SHEET_SESSIONS in wb.sheetnames:
        for row in wb[SHEET_SESSIONS].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            session_id = str(row[0])
            start = datetime.fromisoformat(str(row[1]))
            end_raw = row[2]
            end = datetime.fromisoformat(str(end_raw)) if end_raw else None
            sessions.append(Session(id=session_id, start=start, end=end))
            if end is None:
                active_session_id = session_id

    adjustments: list[HourAdjustment] = []
    if SHEET_ADJUSTMENTS in wb.sheetnames:
        for row in wb[SHEET_ADJUSTMENTS].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            created_raw = row[4] if len(row) > 4 and row[4] else row[1]
            hour_type_raw = row[6] if len(row) > 6 and row[6] else "standard"
            hour_type = str(hour_type_raw).strip().lower()
            if hour_type in {"عادية", "standard"}:
                hour_type = "standard"
            elif hour_type in {"إضافية", "overtime"}:
                hour_type = "overtime"
            else:
                hour_type = "standard"
            adjustments.append(
                HourAdjustment(
                    id=str(row[0]),
                    at=datetime.fromisoformat(str(row[1])),
                    hours_delta=float(row[2]),
                    reason=str(row[3] or ""),
                    created_at=datetime.fromisoformat(str(created_raw)),
                    hour_type=hour_type,
                )
            )

    if SHEET_META in wb.sheetnames:
        meta: dict[str, str] = {}
        for row in wb[SHEET_META].iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                meta[str(row[0])] = "" if row[1] is None else str(row[1])
        if meta.get("active_session_id"):
            active_session_id = meta["active_session_id"] or None

    wb.close()
    return settings, sessions, adjustments, active_session_id


def sync_to_excel_if_needed(
    settings: Settings,
    sessions: list[Session],
    active_session_id: str | None,
    *,
    adjustments: list[HourAdjustment] | None = None,
    force: bool = False,
    today: date | None = None,
) -> Path | None:
    payload = load_payload()
    last_sync_date = payload.get("last_excel_sync_date")

    if not force and not should_sync_today(last_sync_date, today):
        return None

    export_to_excel(settings, sessions, active_session_id, adjustments=adjustments)
    payload["last_excel_sync_date"] = (today or date.today()).isoformat()
    save_data(payload)
    return EXCEL_FILE


def restore_json_from_excel(
    path: Path | None = None,
) -> tuple[Settings, list[Session], list[HourAdjustment], str | None]:
    settings, sessions, adjustments, active_session_id = import_from_excel(path)
    payload = dump_state(
        settings,
        sessions,
        active_session_id,
        adjustments=adjustments,
    )
    payload["last_excel_sync_date"] = date.today().isoformat()
    save_data(payload)
    export_to_excel(settings, sessions, active_session_id, adjustments=adjustments)
    return settings, sessions, adjustments, active_session_id


def export_full_backup(
    settings: Settings,
    sessions: list[Session],
    active_session_id: str | None,
    destination: Path,
    *,
    adjustments: list[HourAdjustment] | None = None,
    json_source: Path | None = None,
) -> tuple[Path, Path | None]:
    """Export a full Excel dump (and optional JSON copy) to a user-chosen location."""
    excel_path = destination
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        excel_path = excel_path.with_suffix(".xlsx")

    written = export_to_excel(
        settings,
        sessions,
        active_session_id,
        adjustments=adjustments,
        path=excel_path,
    )

    json_copy: Path | None = None
    if json_source is not None and json_source.exists():
        json_copy = excel_path.with_suffix(".json")
        shutil.copy2(json_source, json_copy)

    return written, json_copy


def sync_status_text(last_sync_date: str | None) -> str:
    if not last_sync_date:
        return "لم تتم مزامنة Excel بعد"
    if EXCEL_FILE.exists():
        return f"آخر مزامنة Excel: {last_sync_date} — {EXCEL_FILE.name}"
    return f"آخر مزامنة مسجلة: {last_sync_date} (الملف غير موجود)"
